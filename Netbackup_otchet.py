import sys
import pandas as pd 
import subprocess

import re 
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, time
import time
import argparse

def createParser ():
    # Создаем класс парсера
    parser = argparse.ArgumentParser()
    parser.add_argument('-s','--ServerName', required=True )
    parser.add_argument('-a','--Active' , default='Yes')
    return parser

filename = 'Otchet.xlsx'
#wb = openpyxl.load_workbook(filename)
wb = openpyxl.Workbook()
ws = wb.active
row = 2
class Policy:
    def __init__(self,name):
        global servername
        self.name = name
        self.active = "no"
        self.Schedule = []
        self.Residence = []
        self.VolumePool = []
        self.Type = []
        self.Frequency = []
        self.RetentionLevel = []
        self.TimeStart = [[],[],[],[],[],[],[],[]]
        self.TimeEnd = [[],[],[],[],[],[],[],[]]
        self.Client = []
        self.IncrType = []
        print('собираем инфу по политике %s' % self.name)
        policy_info = subprocess.getoutput('Powershell Invoke-Command -ComputerName %s -ScriptBlock {bppllist %s  -L} ' %(servername, self.name) ).split('\n')
        for string in policy_info:
            #print (string)
            if "Policy Type:"  in string : 
                self.PolicyType = (string.split(':')[1].strip()) 
            if "Active:"  in string and 'yes' in string : 
                self.active = 'yes' 
            if "Residence:" in string:
                self.Residence.append(string.split(':')[1].strip())
            if "Volume Pool:" in string: 
                self.VolumePool.append(string.split(':')[1].strip())
            if "Schedule:" in string : 
                self.Schedule.append(string.split(':')[1].strip())
            if "Client/HW/OS/Pri/CIT:" in string :
                self.Client.append(string.split(':')[1].strip())
            if "  Type:" in string : 
                self.Type.append(string)
            if "Frequency:" in string : 
                self.Frequency.append(string.split(':')[1].strip())
            if "Retention Level:" in string : 
                self.RetentionLevel.append(string)
            if "Incr Type:" in string : 
                self.IncrType.append(string)
            if re.search(r'(\d{3}:\d{2}:\d{2}\s{0,}){4}',  string ) :
                arr = re.split(r'\s+',string) 
                ind = len(self.Schedule) - 1
                self.TimeStart[ind].append( int(arr[4].split(':')[0]))
                self.TimeEnd[ind].append( int(arr[5].split(':')[0]))
                

def FillPolicy(pol):
    print('заполняем инфу по политике %s' % pol)
    global row
    
    ws['A%s'% row] = pol.name
    ws['B%s'% row] = pol.PolicyType
    ws['C%s'% row] = pol.active
    ws['D%s'% row] = pol.Residence[0]
    ws['E%s'% row] = '\n '.join(pol.Client)
    ws['F%s'% row] = pol.VolumePool[0]
    if pol.active == 'no':
        return 
    for shed in pol.Schedule:
        row += 1
        index = pol.Schedule.index(shed)
        ws['G%s'% row] = pol.Schedule[index]
        ws['H%s'% row] = pol.Type[index]
        ws['J%s'% row] = pol.RetentionLevel[index]
        ws['K%s'% row] = pol.IncrType[index]
        ws['L%s'% row] = pol.Residence[index+1]
        ws['M%s'% row] = pol.VolumePool[index]
        try:
            ws['I%s'% row] = pol.Frequency[index]
        except:
            pass
        for t in pol.TimeStart[index]: 
            i = pol.TimeStart[index].index(t)
            MarkCells(pol.TimeStart[index][i]+20, pol.TimeEnd[index][i]+20)
        FillJobs(pol.name, pol.Schedule[index])
                
    row += 1
    wb.save(filename)

def MarkCells(TimeStart, TimeEnd,clr='003366FF'):
    Pfill = PatternFill(fill_type='solid', start_color=clr, end_color=clr)
    for i in range(TimeStart,TimeEnd):
        col = get_column_letter(i)
        ws['%s%s'% (col,row)].fill = Pfill
    
'''
def Convert(lst):
        res_dct = {i+1:lst[i]  for i in range(0, len(lst))}
        return res_dct
'''
def CreateXlFile():
    print ('Начато создание файла XL')
    start_time = time.time()
    #wb = openpyxl.Workbook()
    #ws = wb.active
    ws.title = "Netbackup"
    ws.sheet_properties.tabColor = "1072BA"

    # здесь мы меняем цвет решетки в диапазоне. И это происходит дико долго
    def set_border(ws, cell_range):
        thin = Side(border_style="thin", color="9999FF")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    set_border(ws, 't1:t10000') 
    set_border(ws, 'ar1:ar10000') 
    set_border(ws, 'bp1:bp10000')
    set_border(ws, 'cn1:cn10000')  
    set_border(ws, 'dl1:dl10000')
    set_border(ws, 'ej1:ej10000')  
    set_border(ws, 'fh1:fh10000') 
    set_border(ws, 'gf1:gf10000') 
    ws.merge_cells('t1:aq1')
    ws.merge_cells('ar1:bo1')
    ws.merge_cells('bp1:cm1')
    ws.merge_cells('cn1:dk1')
    ws.merge_cells('dl1:ei1')
    ws.merge_cells('ej1:fg1')
    ws.merge_cells('fh1:ge1')
    ws.merge_cells('gf1:hc1')
    ws['T1'] = 'Воскресенье'
    ws['AR1'] = 'Понедельник'
    ws['BP1'] = 'Вторник'
    ws['CN1'] = 'Среда'
    ws['DL1'] = 'Четверг'
    ws['EJ1'] = 'Пятница'
    ws['FH1'] = 'Суббота'
    ws['GF1'] = 'воскресенье'

    ws['A1'] = 'Policy Name'
    ws['B1'] = 'Policy Type'
    ws['C1'] = 'Active'
    ws['D1'] = 'Residence'
    ws['E1'] = 'Client'
    ws['F1'] = 'Volume Pool'
    ws['G1'] = 'Schedule'

    ws['H1'] = 'Type'
    ws['I1'] = 'Frequency'
    ws['J1'] = 'Retention Level'
    ws['K1'] = 'Incr Type'
    ws['L1'] = 'Residence'
    ws['M1'] = 'Volume Pool'

    # Здесь мы меняем размер ячеек чтобы построить решетку для недельки 
    from openpyxl.utils import get_column_letter
    for i in range(20,212):
        column_letter = get_column_letter(i)
        #print(column_letter)
        ws.column_dimensions[column_letter].width = 1
    
    # сохраняем файл и изменения.
    wb.save('Otchet.xlsx')

    print ('закончено содание файла XL')
    print("--- %s seconds ---" % (time.time() - start_time))
    
def ConvertToHourOfWeek(utc):
    d = datetime.fromtimestamp(utc)
    d = d.hour + d.isoweekday()*24
    return d

def FillJobs(name, Sched):
    print ('Заполняем выполненные джобы по политике %s' % name)
    global row
    cols = ['jobid','jobtype','state','status','policy','schedule','client','server',
    'started','elapsed','ended','stunit','tries','operation','kbytes','files','pathlastwritten',
    'percent','jobpid','owner','subtype','policytype','scheduletype','priority','group',
    'masterserver','retentionlevel','retentionperiod','compression','kbytestobewritten',
    'filestobe written','filelistcount','[files]','parentjob','kbpersec','copy','robot',
    'vault','profile','session','ejecttapes','srcstunit','srcserver','srcmedia','dstmedia',
    'stream','suspendable','resumable','restartable','datamovement','snapshot','backupid',
    'killable','controllinghost','offhosttype','ftusage','queuereason','dedupratio',
    'acceleratorOptimization','instancedbname']
    #Читаем джобы из файла, удобно для отладки, поэтому сделал через временный файл.
    #df = pd.read_csv("bpdbjobs1",encoding = 'ISO-8859-1',on_bad_lines='skip',names=cols )
    df = pd.read_csv("Joblist.csv",on_bad_lines='skip',index_col=False, names=cols )
    jobs = df[df.policy =='%s'% name ][['jobid','status','policy','schedule','started','elapsed','ended','stunit']].copy()
    for val in jobs.values:
        if val[3] == Sched:
            row += 1
            ws['N%s'% row] = val[0]
            ws['O%s'% row] = val[1]
            ws['P%s'% row] = val[3]
            #ws['S%s'% row] = val[7]
            ws['Q%s'% row] = datetime.fromtimestamp(val[4])
            ws['R%s'% row] = datetime.fromtimestamp(val[6])
            TimeStart = ConvertToHourOfWeek(val[4])
            TimeEnd = ConvertToHourOfWeek(val[6])
            if TimeStart > 24*7:
                if TimeEnd >= TimeStart:
                    TimeEnd = TimeEnd - 24*7
                TimeStart = TimeStart - 24*7
            if TimeEnd == TimeStart:
                TimeEnd+=1
            if val[1]==0:
                MarkCells(TimeStart+20, TimeEnd+20,clr='0033CCCC')
            else:
                MarkCells(TimeStart+20, TimeEnd+20,clr='00FF00FF')


    
    #print (df)
    #df[df['policy']=='NBP01-BMR-KSDD' ] ['schedule']]
    #df[df.policy =='NBP01-BMR-KSDD' ][['jobid','policy','schedule','started','elapsed','ended']]
    #tab = df[df.policy =='NBP01-BMR-KSDD' ][['jobid','policy','schedule','started','elapsed','ended']].copy()
    #tab.iloc[1]['policy']
def main():
    
    parser = createParser()
    namespace = parser.parse_args(sys.argv[1:])
    
    print (namespace)
    global activepolicy
    activepolicy = namespace.Active
    global servername
    servername = namespace.ServerName
    
    

    # Получаем данные запросами с сервака. 
    # Данные о политиках которые существуют активных и неактивных
    print('собираем инфу о политиках')
    ListOfPolicyNames = subprocess.getoutput('Powershell Invoke-Command -ComputerName %s -ScriptBlock {bppllist -l} '% servername).split()
    
    # Данные о выполненных и выполняющихся джобах. и загружаем их в файл.  
    # bpdbjobs -gdm -file C:\temp\bpdbjobs1
    with open('Joblist.csv', "w") as outfile:
        subprocess.run('Powershell Invoke-Command -ComputerName %s -ScriptBlock {bpdbjobs -gdm } '% servername, stdout=outfile )

    # здесь создаем файл эксель и рисуем в нем предварительные красивости.
    CreateXlFile()
    

    ########## 1 здесь пока ставим заглушку и тренируемся на одной политике. 
    
    for PolicyName in ListOfPolicyNames:
        #print(PolicyName)
        PolicyName = Policy(PolicyName)
        FillPolicy(PolicyName)
    
    #PolicyName = 'NBP01-BMR-KSDD'
    #pol = Policy(PolicyName)
    #FillPolicy(pol)
    


if __name__ == "__main__":
    main()